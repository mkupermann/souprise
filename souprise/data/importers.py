"""Importers that turn external data into indexable entries.

Every importer returns the same shape the retrievers expect: a list of
dicts with 'id', 'text', and 'metadata'. Records are rendered as
"Column: value" lines, matching the format of the synthetic generators,
so fine-tuned models see familiar structure.

Supported sources: CSV (stdlib), Excel (openpyxl via the 'excel' extra),
JSONL, and PostgreSQL (sqlalchemy via the 'postgres' extra).

License: Apache-2.0
Copyright 2026 Michael Kupermann
"""

import csv
import json
from typing import Any, Dict, Iterable, List, Optional

Entry = Dict[str, Any]


def entries_from_records(
    records: Iterable[Dict[str, Any]],
    id_column: Optional[str] = None,
    text_columns: Optional[List[str]] = None,
    tag_columns: Optional[List[str]] = None,
) -> List[Entry]:
    """Convert tabular records into indexable entries.

    Args:
        records: Iterable of dicts, one per row.
        id_column: Column used as the entry id. Falls back to row_<n>.
        text_columns: Columns rendered into the searchable text.
            Defaults to every column except the id column.
        tag_columns: Columns whose values become metadata tags.

    Returns:
        List of {'id', 'text', 'metadata'} dicts.
    """
    entries: List[Entry] = []
    for i, record in enumerate(records):
        entry_id = str(record.get(id_column, f"row_{i}")) if id_column else f"row_{i}"
        columns = text_columns or [c for c in record.keys() if c != id_column]
        text = "\n".join(
            f"{column}: {record[column]}"
            for column in columns
            if record.get(column) not in (None, "")
        )
        tags = [
            str(record[column])
            for column in (tag_columns or [])
            if record.get(column) not in (None, "")
        ]
        entries.append({
            "id": entry_id,
            "text": text if not id_column else f"{entry_id}\n{text}",
            "metadata": {"tags": tags} if tags else {},
        })
    return entries


def load_csv(
    path: str,
    id_column: Optional[str] = None,
    text_columns: Optional[List[str]] = None,
    tag_columns: Optional[List[str]] = None,
    delimiter: Optional[str] = None,
) -> List[Entry]:
    """Load entries from a CSV file. The delimiter is sniffed unless given."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        if delimiter is None:
            sample = f.read(4096)
            f.seek(0)
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                delimiter = ","
        reader = csv.DictReader(f, delimiter=delimiter)
        return entries_from_records(reader, id_column, text_columns, tag_columns)


def load_excel(
    path: str,
    sheet: Optional[str] = None,
    id_column: Optional[str] = None,
    text_columns: Optional[List[str]] = None,
    tag_columns: Optional[List[str]] = None,
) -> List[Entry]:
    """Load entries from an Excel workbook. First row is the header."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            'openpyxl not installed. Install with: pip install -e ".[excel]"'
        ) from e

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows))]
    records = [dict(zip(header, row)) for row in rows]
    workbook.close()
    return entries_from_records(records, id_column, text_columns, tag_columns)


def load_jsonl(path: str) -> List[Entry]:
    """Load entries from JSONL.

    Lines already shaped as {'id', 'text', ...} pass through; any other
    flat object is rendered like a tabular record.
    """
    entries: List[Entry] = []
    with open(path, encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            if "text" in record:
                entries.append({
                    "id": str(record.get("id", f"row_{i}")),
                    "text": record["text"],
                    "metadata": record.get("metadata") or {},
                })
            else:
                entries.extend(entries_from_records([record]))
                entries[-1]["id"] = str(record.get("id", f"row_{i}"))
    return entries


def load_postgres(
    dsn: str,
    query: str,
    id_column: Optional[str] = None,
    text_columns: Optional[List[str]] = None,
    tag_columns: Optional[List[str]] = None,
) -> List[Entry]:
    """Load entries from a PostgreSQL query.

    Args:
        dsn: SQLAlchemy connection string, e.g.
            postgresql://user@localhost:5432/mydb
        query: SELECT statement; column names become the record keys.
    """
    try:
        from sqlalchemy import create_engine, text
    except ImportError as e:
        raise ImportError(
            'sqlalchemy not installed. Install with: pip install -e ".[postgres]"'
        ) from e

    engine = create_engine(dsn)
    try:
        with engine.connect() as connection:
            result = connection.execute(text(query))
            records = [dict(row._mapping) for row in result]
    finally:
        engine.dispose()
    return entries_from_records(records, id_column, text_columns, tag_columns)
